#!/usr/bin/env python3
"""生成长江证券模拟交割单 Excel 文件（脱敏数据）"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime, timedelta
import random

# 模拟股票池（脱敏，用常见ETF代替）
STOCKS = [
    ("510300", "沪深300ETF"),
    ("510500", "中证500ETF"),
    ("159915", "创业板ETF"),
    ("588000", "科创50ETF"),
    ("512000", "券商ETF"),
    ("512690", "酒ETF"),
    ("515700", "新能源ETF"),
    ("518880", "黄金ETF"),
    ("000001", "平安银行"),
    ("600036", "招商银行"),
]

def generate_mock_trades():
    """生成模拟交易记录"""
    trades = []
    start_date = datetime(2026, 3, 1)
    
    for i in range(20):
        trade_date = start_date + timedelta(days=random.randint(0, 60))
        stock_code, stock_name = random.choice(STOCKS)
        direction = random.choice(["买入", "卖出"])
        
        # 价格范围根据股票类型调整
        if "ETF" in stock_name:
            price = round(random.uniform(0.5, 5.0), 3)
        else:
            price = round(random.uniform(10.0, 50.0), 2)
        
        quantity = random.choice([100, 200, 300, 500, 1000])
        amount = round(price * quantity, 2)
        
        # 手续费按万分之三估算
        commission = round(amount * 0.0003, 2)
        stamp_tax = round(amount * 0.001, 2) if direction == "卖出" else 0
        transfer_fee = round(amount * 0.00002, 2)
        
        # 发生金额（买入为负，卖出为正）
        if direction == "买入":
            total = -(amount + commission + stamp_tax + transfer_fee)
        else:
            total = amount - commission - stamp_tax - transfer_fee
        
        trades.append({
            "成交日期": trade_date.strftime("%Y-%m-%d"),
            "成交时间": f"{random.randint(9, 14)}:{random.randint(30, 59)}:{random.randint(0, 59)}",
            "证券代码": stock_code,
            "证券名称": stock_name,
            "操作": direction,
            "成交数量": quantity,
            "成交价格": price,
            "成交金额": amount,
            "手续费": commission,
            "印花税": stamp_tax,
            "过户费": transfer_fee,
            "佣金": commission,
            "发生金额": round(total, 2),
            "成交编号": f"CJ{trade_date.strftime('%Y%m%d')}{random.randint(100000, 999999)}",
            "股东代码": "A123456789",
            "交易市场": "上海A股" if stock_code.startswith(("51", "60")) else "深圳A股",
            "备注": "",
        })
    
    # 按日期排序
    trades.sort(key=lambda x: x["成交日期"])
    return trades

def create_excel(filepath):
    """创建 Excel 文件"""
    trades = generate_mock_trades()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "交割单"
    
    # 标题行样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 写入标题
    headers = list(trades[0].keys())
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # 写入数据
    for row_idx, trade in enumerate(trades, 2):
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=trade[header])
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center' if col_idx <= 5 else 'right')
            
            # 买入红色，卖出绿色（中国习惯）
            if header == "操作":
                if trade[header] == "买入":
                    cell.font = Font(color="FF0000", bold=True)
                else:
                    cell.font = Font(color="008000", bold=True)
            
            # 发生金额为负标红
            if header == "发生金额" and trade[header] < 0:
                cell.font = Font(color="FF0000")
    
    # 调整列宽
    col_widths = [12, 10, 10, 14, 8, 10, 10, 12, 10, 10, 10, 10, 12, 20, 12, 10, 10]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    
    # 添加汇总行
    summary_row = len(trades) + 2
    ws.cell(row=summary_row, column=1, value="汇总信息").font = Font(bold=True, size=12)
    ws.cell(row=summary_row + 1, column=1, value=f"总成交笔数: {len(trades)}")
    ws.cell(row=summary_row + 2, column=1, value=f"买入笔数: {sum(1 for t in trades if t['操作']=='买入')}")
    ws.cell(row=summary_row + 3, column=1, value=f"卖出笔数: {sum(1 for t in trades if t['操作']=='卖出')}")
    
    wb.save(filepath)
    print(f"已生成模拟文件: {filepath}")
    print(f"共 {len(trades)} 笔交易记录")
    return filepath

if __name__ == "__main__":
    filepath = "/Users/leijiang/WorkBuddy/moneybag-for-claudecode/tests/fixtures/broker_statements/changjiang_delivery_mock.xlsx"
    create_excel(filepath)
